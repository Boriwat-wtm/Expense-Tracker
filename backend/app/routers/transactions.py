import io
from datetime import date
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import extract
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import get_current_user
from ..models.transaction import Transaction, TransactionSource, TransactionType
from ..models.user import User
from ..schemas.transaction import TransactionCreate, TransactionOut, TransactionUpdate

router = APIRouter(prefix="/transactions", tags=["transactions"])


@router.get("", response_model=List[TransactionOut])
def list_transactions(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    type: Optional[TransactionType] = None,
    source: Optional[TransactionSource] = None,
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, le=500),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if start_date:
        q = q.filter(Transaction.date >= start_date)
    if end_date:
        q = q.filter(Transaction.date <= end_date)
    if type:
        q = q.filter(Transaction.type == type)
    if source:
        q = q.filter(Transaction.source == source)
    return q.order_by(Transaction.date.desc()).offset(skip).limit(limit).all()


@router.post("", response_model=TransactionOut, status_code=201)
def create_transaction(
    data: TransactionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    txn = Transaction(**data.model_dump(), user_id=current_user.id)
    db.add(txn)
    db.commit()
    db.refresh(txn)
    return txn


@router.put("/{txn_id}", response_model=TransactionOut)
def update_transaction(
    txn_id: UUID,
    data: TransactionUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    txn = (
        db.query(Transaction)
        .filter(Transaction.id == txn_id, Transaction.user_id == current_user.id)
        .first()
    )
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(txn, field, value)
    db.commit()
    db.refresh(txn)
    return txn


@router.delete("/{txn_id}", status_code=204)
def delete_transaction(
    txn_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    txn = (
        db.query(Transaction)
        .filter(Transaction.id == txn_id, Transaction.user_id == current_user.id)
        .first()
    )
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    db.delete(txn)
    db.commit()


@router.get("/export")
def export_transactions(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    type: Optional[TransactionType] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    q = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if start_date:
        q = q.filter(Transaction.date >= start_date)
    if end_date:
        q = q.filter(Transaction.date <= end_date)
    if type:
        q = q.filter(Transaction.type == type)
    if month:
        q = q.filter(extract("month", Transaction.date) == month)
    if year:
        q = q.filter(extract("year", Transaction.date) == year)
    rows = q.order_by(Transaction.date.asc(), Transaction.created_at.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "รายการธุรกรรม"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    income_font = Font(color="15803D", bold=True)
    expense_font = Font(color="DC2626", bold=True)
    total_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # ── Header row ───────────────────────────────────────────────────────────
    headers = ["ลำดับ", "วันที่", "คำอธิบาย", "หมวดหมู่", "ประเภท", "Source", "ยอดเงิน (บาท)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border

    # ── Data rows ────────────────────────────────────────────────────────────
    TYPE_TH = {"income": "รายรับ", "expense": "รายจ่าย"}
    SOURCE_TH = {"slip": "สลิป", "pdf": "PDF", "manual": "Manual"}
    total_income = 0.0
    total_expense = 0.0

    for i, txn in enumerate(rows, 1):
        amount = float(txn.amount)
        r = i + 1
        ws.cell(row=r, column=1, value=i).alignment = center
        ws.cell(row=r, column=2, value=str(txn.date))
        ws.cell(row=r, column=3, value=txn.description or "")
        ws.cell(row=r, column=4, value=txn.category or "")
        ws.cell(row=r, column=5, value=TYPE_TH.get(txn.type.value, txn.type.value))
        ws.cell(row=r, column=6, value=SOURCE_TH.get(txn.source.value, txn.source.value))
        amount_cell = ws.cell(row=r, column=7, value=amount)
        amount_cell.number_format = '#,##0.00'
        # Color
        f = income_font if txn.type.value == "income" else expense_font
        amount_cell.font = f
        ws.cell(row=r, column=5).font = f
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = cell_border
        if txn.type.value == "income":
            total_income += amount
        else:
            total_expense += amount

    # ── Summary rows ─────────────────────────────────────────────────────────
    last_data = len(rows) + 1
    summary_data = [
        ("รายรับรวม", total_income, "15803D"),
        ("รายจ่ายรวม", total_expense, "DC2626"),
        ("ยอดคงเหลือ", total_income - total_expense, "1E40AF"),
    ]
    for offset, (label, val, color) in enumerate(summary_data):
        r = last_data + 1 + offset
        merged = ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        merged.font = Font(bold=True, color=color, size=11)
        merged.fill = total_fill
        merged.alignment = Alignment(horizontal="right", vertical="center")
        amt = ws.cell(row=r, column=7, value=val)
        amt.number_format = '#,##0.00'
        amt.font = Font(bold=True, color=color, size=11)
        amt.fill = total_fill
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = cell_border

    # ── Column widths ────────────────────────────────────────────────────────
    widths = [8, 14, 40, 20, 12, 10, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    # ── Stream back ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"transactions_{current_user.username}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
